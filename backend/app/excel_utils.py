from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

DATA_DIR = Path(__file__).resolve().parents[1] / "data"
WORKBOOK_PATH = DATA_DIR / "submissions.xlsx"


def _get_or_create_workbook(sheet_name: str, headers: list[str]):
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if WORKBOOK_PATH.exists():
        workbook = load_workbook(WORKBOOK_PATH)
    else:
        workbook = Workbook()

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(headers)

    # If default sheet exists and is unused, remove it.
    if "Sheet" in workbook.sheetnames and len(workbook["Sheet"]["A"]) == 1 and workbook["Sheet"]["A1"].value is None:
        workbook.remove(workbook["Sheet"])

    return workbook, sheet


def _append_row(sheet_name: str, headers: list[str], row: list[str | int]):
    workbook, sheet = _get_or_create_workbook(sheet_name, headers)
    sheet.append(row)
    workbook.save(WORKBOOK_PATH)


def append_enquiry_record(name: str, email: str, service: str | None, message: str) -> None:
    _append_row(
        sheet_name="enquiries",
        headers=["Timestamp", "Name", "Email", "Service", "Message"],
        row=[
            datetime.now().isoformat(timespec="seconds"),
            name,
            email,
            service or "Not specified",
            message,
        ],
    )


def append_feedback_record(name: str, email: str, rating: int, message: str) -> None:
    _append_row(
        sheet_name="feedbacks",
        headers=["Timestamp", "Name", "Email", "Rating", "Message"],
        row=[
            datetime.now().isoformat(timespec="seconds"),
            name,
            email,
            rating,
            message,
        ],
    )


def append_career_record(
    name: str,
    email: str,
    phone: str,
    position: str,
    resume_url: str,
) -> None:
    _append_row(
        sheet_name="careers",
        headers=["Timestamp", "Name", "Email", "Phone", "Position", "Resume URL"],
        row=[
            datetime.now().isoformat(timespec="seconds"),
            name,
            email,
            phone,
            position,
            resume_url,
        ],
    )
